"""Patch booking sheet Column C/I/J only (preserve date cell types).

SOURCE (only):
  alan-shared-resources/csv/Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm

NOTE: openpyxl strips Excel x14:dataValidations (column I ProductList dropdown).
After save, re-inject via restore-booking-column-i-validation.py in shared-resources.
"""
from __future__ import annotations

import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

SOURCE = Path(
    r"G:/Dropbox/alan ranger photography/Website Code/alan-shared-resources/csv/Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm"
)

SENSOR_LANDING = "https://www.alanranger.com/photography-services-near-me/camera-sensor-clean/"
GUEST_LANDING = "https://www.alanranger.com/professional-commercial-photographer-coventry"
PRINT_LANDING = "https://www.alanranger.com/fine-art-prints"
KASE_LANDING = "https://www.alanranger.com/photography-workshops-near-me"
ACADEMY_LANDING = "https://www.alanranger.com/free-online-photography-course"

SENSOR_EVENTS = {"sensor clean", "sensore clean", "sensor clean x 3"}
PRINT_EVENTS = {"tripod", "pocket guides"}
KASE_EVENTS = {"kase affiliates", "kase royalties"}

DATE_STRING_RE = re.compile(
    r"^(?:Mon|Tue|Wed|Thu|Fri|Sat|Sun)\s+"
    r"(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+"
    r"\d{1,2}\s+\d{4}"
)


def norm_event(value) -> str:
    return str(value or "").strip().lower()


def repair_date_cell(ws, row: int) -> bool:
    val = ws.cell(row, 1).value
    if not isinstance(val, str):
        return False
    if not DATE_STRING_RE.match(val.strip()):
        return False
    try:
        core = val.split(" GMT")[0].strip()
        ws.cell(row, 1).value = datetime.strptime(core, "%a %b %d %Y %H:%M:%S")
        return True
    except ValueError:
        return False


def apply_rules(ws, row: int) -> str | None:
    ev = norm_event(ws.cell(row, 6).value)
    if ev in SENSOR_EVENTS:
        ws.cell(row, 3).value = "11 Commissions"
        ws.cell(row, 9).value = "Sensor Clean Service (historical)"
        ws.cell(row, 10).value = SENSOR_LANDING
        return f"sensor:{ev}"
    if ev == "guest blog":
        ws.cell(row, 3).value = "11 Commissions"
        ws.cell(row, 9).value = "Commission - Editorial / Guest Blog / Judging"
        ws.cell(row, 10).value = GUEST_LANDING
        return "guest blog"
    if ev in PRINT_EVENTS:
        ws.cell(row, 3).value = "10. Prints & Royalties"
        ws.cell(row, 9).value = "Print Sale - Generic (historical)"
        ws.cell(row, 10).value = PRINT_LANDING
        return f"print-merch:{ev}"
    if ev in KASE_EVENTS:
        ws.cell(row, 3).value = "10. Prints & Royalties"
        ws.cell(row, 9).value = "Royalties & Affiliate Income"
        ws.cell(row, 10).value = KASE_LANDING
        return f"kase:{ev}"
    if ev == "e-book foundation":
        cat = str(ws.cell(row, 3).value or "").strip()
        if cat != "12. Other":
            ws.cell(row, 3).value = "12. Academy"
        ws.cell(row, 9).value = "Academy - Membership & Income"
        ws.cell(row, 10).value = ACADEMY_LANDING
        return "e-book foundation"
    return None


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, keep_vba=True, data_only=False)
    patched = 0
    dates_repaired = 0
    by_sheet: dict[str, int] = {}

    for name in wb.sheetnames:
        m = re.match(r"^Sales\s+(\d{4})$", name, re.I)
        if not m or int(m.group(1)) < 2024:
            continue
        ws = wb[name]
        header_row = None
        for r in range(100, 251):
            if str(ws.cell(r, 1).value or "").strip().lower() == "date":
                header_row = r
                break
        if not header_row:
            continue
        blank_run = 0
        sheet_count = 0
        for r in range(header_row + 1, ws.max_row + 1):
            date_val = ws.cell(r, 1).value
            if date_val is None or str(date_val).strip() == "":
                blank_run += 1
                if blank_run >= 50:
                    break
                continue
            blank_run = 0
            if repair_date_cell(ws, r):
                dates_repaired += 1
            change = apply_rules(ws, r)
            if change:
                sheet_count += 1
                patched += 1
        if sheet_count:
            by_sheet[name] = sheet_count

    wb.save(SOURCE)
    print(f"patched rows: {patched} {by_sheet}")
    print(f"dates repaired: {dates_repaired}")

    restore = Path(
        r"G:/Dropbox/alan ranger photography/Website Code/alan-shared-resources/scripts/"
        r"restore-booking-column-i-validation.py"
    )
    if restore.is_file():
        print("Restoring column I ProductList validation (openpyxl strips x14 DV)...")
        subprocess.run([sys.executable, str(restore)], check=True)
    else:
        print("WARN: restore-booking-column-i-validation.py not found — re-run manually")


if __name__ == "__main__":
    main()
