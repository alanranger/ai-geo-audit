"""Recalculate all formulas in the live booking sheet and save cached values.

Required on Windows with Microsoft Excel installed (uses COM via pywin32).

SOURCE (only):
  alan-shared-resources/csv/Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm

Usage:
  pip install pywin32 openpyxl
  python scripts/recalc-booking-sheet-xlsm.py
  python scripts/recalc-booking-sheet-xlsm.py --verify-only
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

SOURCE = Path(
    r"G:/Dropbox/alan ranger photography/Website Code/alan-shared-resources/csv/"
    r"Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm"
)
SALES_TABS = ("Sales 2024", "Sales 2025", "Sales 2026")


def verify_cached_values(path: Path) -> list[str]:
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(path, data_only=True)
    issues: list[str] = []
    for name in SALES_TABS:
        if name not in wb.sheetnames:
            issues.append(f"{name}: tab missing")
            continue
        ws = wb[name]
        missing = [
            f"{get_column_letter(c)}{r}"
            for r in range(6, 18)
            for c in range(11, 23)
            if ws.cell(r, c).value is None
        ]
        ytd_row = 46 if name == "Sales 2024" else 47 if name == "Sales 2025" else 48
        ytd = ws.cell(ytd_row, 10).value
        if missing:
            issues.append(f"{name}: {len(missing)} monthly cells without cached value")
        if ytd is None:
            issues.append(f"{name}: YTD Actual (J{ytd_row}) has no cached value")
    return issues


def recalc_with_excel(path: Path) -> None:
    import win32com.client

    if not path.is_file():
        raise FileNotFoundError(path)
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(path.resolve()))
        for tab in SALES_TABS:
            wb.Worksheets(tab).Activate()
        excel.CalculateFullRebuild()
        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        excel.Quit()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        type=Path,
        default=SOURCE,
        help="Path to Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm",
    )
    parser.add_argument(
        "--verify-only",
        action="store_true",
        help="Check cached formula values without opening Excel",
    )
    args = parser.parse_args()
    path = args.source.resolve()
    print("FILE:", path)

    if args.verify_only:
        issues = verify_cached_values(path)
        if issues:
            print("ISSUES:")
            for item in issues:
                print(" ", item)
            return 1
        print("OK: monthly grid + YTD cells have cached values")
        return 0

    recalc_with_excel(path)
    issues = verify_cached_values(path)
    if issues:
        print("Recalc finished but verification failed:")
        for item in issues:
            print(" ", item)
        return 1
    print("OK: Excel recalc saved; monthly grid + YTD cached values present")
    return 0


if __name__ == "__main__":
    sys.exit(main())
